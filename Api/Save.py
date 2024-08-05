"""module contained funcs for save city, address and product to DB."""

from io import BytesIO

import openpyxl

from aiogram_dialog import DialogManager

from Api.DB import (
    get_city, add_city,
    get_addres, add_address,
    add_product
)

from Models.product import Product


async def save_product(manager: DialogManager) -> None:
    """Get needle data from dialog_data and call save_all."""
    city = manager.dialog_data.get("city", "")
    address = manager.dialog_data.get("address", "")
    product_name = manager.dialog_data.get("product_name", "")
    product_descriptions = manager.dialog_data.get("product_description", "")
    product_price = manager.dialog_data.get("product_price", "")
    product = Product(product_name, product_descriptions, product_price)
    await save_all(city, address, product)


async def save_all(city: str, address: str, product: Product) -> None:
    """Get city, address and product and after save theirs to db."""
    city_id = await get_city(city)
    if city_id is None:
        city_id = await add_city(city)

    address_id = await get_addres(address, city_id)
    if address_id is None:
        address_id = await add_address(address, city_id)

    await add_product(product, address_id)


async def save_with_xlsx(xlsx_bytes: BytesIO) -> None:
    """Get data from xlsx file."""
    table = openpyxl.load_workbook(xlsx_bytes, read_only=True).active

    row = 2
    while table.cell(row, 1).value:
        print(table.cell(row, 1).value)
        city = table.cell(row, 1).value
        address = table.cell(row, 2).value
        product_name = table.cell(row, 3).value
        product_discription = table.cell(row, 4).value
        product_price = table.cell(row, 5).value
        product = Product(product_name, product_discription, product_price)
        await save_all(city, address, product)
        row += 1
